"""
DIOPTRA FIDC — Dashboard de Governança para FIDCs de Duplicatas/PME
Autor: Eduardo Fochesatto
Licença: MIT
"""
import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border
from openpyxl.utils.dataframe import data_frame_to_rows

# ── CONFIGURAÇÃO ──
NOME_FERRAMENTA = "Dioptra FIDC"
AUTOR = "Eduardo Fochesatto"
REPOSITORIO = "https://github.com/eduardofochesatto/dioptra-fidc"
VERSAO = "Julho/2026"

# ── Cores ──
AZUL_ESCURO = "1F3864"
AZUL_MEDIO  = "2E75B6"
BRANCO      = "FFFFFF"
VERDE       = "C6EFCE"
VERMELHO    = "FFC7CE"
AMARELO     = "FFEB9C"
CINZA_L     = "F5F5F5"
CINZA_T     = "808080"
PRETO       = "1A1A1A"

FILL_H   = PatternFill("solid", fgColor=AZUL_ESCURO)
FILL_Z   = PatternFill("solid", fgColor=CINZA_L)
FILL_V   = PatternFill("solid", fgColor=VERDE)
FILL_R   = PatternFill("solid", fgColor=VERMELHO)
FILL_A   = PatternFill("solid", fgColor=AMARELO)

FONT_H  = Font(name="Calibri", bold=True, color=BRANCO, size=11)
FONT_T  = Font(name="Calibri", bold=True, size=18, color=AZUL_ESCURO)
FONT_T2 = Font(name="Calibri", bold=True, size=14, color=AZUL_ESCURO)
FONT_S  = Font(name="Calibri", italic=True, size=11, color=CINZA_T)
FONT_D  = Font(name="Calibri", bold=True, size=12, color=AZUL_ESCURO)
FONT_N  = Font(name="Calibri", size=10)
FONT_A  = Font(name="Calibri", bold=True, size=10, color="CC0000")

AL   = Alignment(horizontal="left", vertical="top", wrap_text=True)
AR   = Alignment(horizontal="right", vertical="center")
AC   = Alignment(horizontal="center", vertical="center")

# ── Benchmarks ──
BENCH = {
    "CDI": {"2022":12.40,"2023":13.05,"2024":10.80,"2025":11.20,"2026":10.50},
    "IMAB":{"2022":10.10,"2023":15.80,"2024":8.40,"2025":12.50,"2026":11.80},
    "IPCA":{"2022":5.79,"2023":4.62,"2024":4.30,"2025":4.80,"2026":4.50},
}

def _cab(ws, l, n):
    for c in range(1, n+1):
        cl = ws.cell(row=l, column=c)
        cl.fill = FILL_H; cl.font = FONT_H; cl.alignment = AC

def _zeb(ws, i, f, n):
    for r in range(i, f+1):
        if (r-i) % 2 == 1:
            for c in range(1, n+1):
                ws.cell(row=r, column=c).fill = FILL_Z

def _assin(ws, l):
    ws.cell(row=l, column=1, value=f"{NOME_FERRAMENTA} — {AUTOR} — {REPOSITORIO}").font = FONT_S

def criar_capa(wb):
    ws = wb.active; ws.title = "0. Capa"
    itens = [
        (1, f"{NOME_FERRAMENTA}", FONT_T),
        (2, "Dashboard de Governança para FIDCs de Duplicatas/PME", FONT_D),
        (3, f"{AUTOR} — {VERSAO}", FONT_S),
        (5, "╔╗", None),
        (6, "║               AVISO IMPORTANTE                         ║", None),
        (7, "╠╣", None),
        (8, "║  Este dashboard é um esforço VOLUNTÁRIO de transparência ║", None),
        (9, "║  baseado em dados públicos autodeclarados à CVM.        ║", None),
        (10,"║                                                        ║", None),
        (11,"║  ⚠  LIMITAÇÕES:                                       ║", None),
        (12,"║  1. DADOS AUTODECLARADOS — sem auditoria independente  ║", None),
        (13,"║  2. DEFASAGEM DE 30-45 DIAS — você olha para o passado║", None),
        (14,"║  3. PREENCHIMENTO PARCIAL — PDD e recompra incompletos ║", None),
        (15,"║  4. NÃO SUBSTITUI DILIGÊNCIA — triagem, não decisão   ║", None),
        (16,"║  5. CLASSIFICAÇÃO Duplicatas/PME é aproximada         ║", None),
        (17,"║                                                        ║", None),
        (18,"║  NENHUMA DECISÃO DE INVESTIMENTO DEVE SER TOMADA      ║", None),
        (19,"║  EXCLUSIVAMENTE COM BASE NESTES NÚMEROS.              ║", None),
        (20,"╚╝", None),
        (22, f"📬  {AUTOR}", FONT_D),
        (23, f"Repositório: {REPOSITORIO}", FONT_S),
        (24, "Licença: MIT", FONT_S),
        (26, "📋  Abas:", FONT_D),
        (27, "1. Ranking Completo — filtro automático por coluna", FONT_N),
        (28, "2. Top 10 por Métrica — melhor fundo em cada item", FONT_N),
        (29, "3. CDI/IMAB — comparação de rentabilidade com benchmarks", FONT_N),
        (30, "4. Governança — estatísticas descritivas + interpretação", FONT_N),
        (31, "5. Propósito — ensaio reflexivo", FONT_N),
        (32, "6. Fontes — documentação completa dos dados", FONT_N),
        (33, "7. Glossário — definição de cada métrica", FONT_N),
    ]
    for l, txt, st in itens:
        c = ws.cell(row=l, column=1, value=txt)
        if st: c.font = st
    for l in [6, 11, 18, 19]:
        ws.cell(row=l, column=1).font = Font(name="Consolas", bold=True, size=10, color="CC0000")
    for l in range(5, 21):
        if ws.cell(row=l, column=1).font == Font():
            ws.cell(row=l, column=1).font = Font(name="Consolas", size=10, color=PRETO)
    ws.column_dimensions["A"].width = 120
    ws.sheet_properties.tabColor = "CC0000"

def criar_ranking(wb, df):
    ws = wb.create_sheet("1. Ranking Completo")
    MAPA = {'CNPJ_FUNDO':'CNPJ do Fundo','DENOMINACAO_SOCIAL':'Nome do Fundo',
            'GESTORA':'Gestora','CLASSE':'Classe','VL_PL':'PL (R$ milhões)',
            'PRAZO_MEDIO':'Prazo Médio (dias)','RENTABILIDADE':'Rentabilidade (% a.a.)',
            'PDD_PCT':'PDD (% PL)','RECOMPRA_PCT':'Recompra (%)',
            'PCT_SUBORDINADO':'Subordinado (%)','OVERCOLLATERALIZATION':'Overcollateralization (x)',
            'NUM_SACADOS':'Nº Sacados','NUM_CEDENTES':'Nº Cedentes',
            'CONC_TOP5_SACADOS':'Conc. Top 5 Sacados (%)',
            'CONC_TOP5_CEDENTES':'Conc. Top 5 Cedentes (%)',
            'PCT_VENCIDOS':'Vencidos +90d (%)','PCT_LIQUIDEZ':'Liquidez (%)',
            'TAXA_ADM':'Taxa Adm (% a.a.)','RATING':'Rating'}
    cols = [c for c in MAPA if c in df.columns]
    dx = df[cols].copy(); dx.columns = [MAPA[c] for c in cols]
    if 'PL (R$ milhões)' in dx.columns:
        dx['PL (R$ milhões)'] = (dx['PL (R$ milhões)']/1_000_000).round(2)
    dx = dx.sort_values('PL (R$ milhões)', ascending=False)
    for r, row in enumerate(data_frame_to_rows(dx, index=False, header=True), 1):
        for c, v in enumerate(row, 1): ws.cell(row=r, column=c, value=v)
    nc = len(dx.columns); ul = len(dx)+1
    _cab(ws, 1, nc); _zeb(ws, 2, ul, nc); _assin(ws, ul+2)
    ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=nc).column_letter}{ul}"
    ws.freeze_panes = "A2"
    for ci, col in enumerate(dx.columns, 1):
        comp = max(len(str(col)), dx[col].astype(str).str.len().max() if len(dx)>0 else 0)
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = min(comp+3, 32)

def criar_top10(wb, df):
    ws = wb.create_sheet("2. Top 10 por Métrica")
    METS = [("PL (R$ milhões)","VL_PL",False),("Rentabilidade (% a.a.)","RENTABILIDADE",False),
            ("Menor PDD (% PL)","PDD_PCT",True),("Maior Overcollateralization","OVERCOLLATERALIZATION",False),
            ("Maior Liquidez (%)","PCT_LIQUIDEZ",False),("Menor Conc. Top 5 Sacados (%)","CONC_TOP5_SACADOS",True),
            ("Menor Conc. Top 5 Cedentes (%)","CONC_TOP5_CEDENTES",True),("Maior Nº de Sacados","NUM_SACADOS",False),
            ("Menor % Vencidos +90d","PCT_VENCIDOS",True),("Maior Recompra (%)","RECOMPRA_PCT",False)]
    linha = 1
    for tit, col, asc in METS:
        if col not in df.columns: continue
        t10 = df.sort_values(col, ascending=asc).head(10)
        ws.cell(row=linha, column=1, value=f"TOP 10 — {tit}").font = FONT_D; linha += 1
        for c, t in enumerate(["#","Fundo","Gestora","Valor"], 1): ws.cell(row=linha, column=c, value=t)
        _cab(ws, linha, 4); linha += 1
        for idx, (_, r) in enumerate(t10.iterrows(), 1):
            ws.cell(row=linha, column=1, value=idx).alignment = AC
            ws.cell(row=linha, column=2, value=str(r.get('DENOMINACAO_SOCIAL',''))[:60])
            ws.cell(row=linha, column=3, value=r.get('GESTORA',''))
            v = r[col]
            if col == 'VL_PL': v = round(v/1_000_000, 2)
            ws.cell(row=linha, column=4, value=round(v,2)).alignment = AR; linha += 1
        linha += 2
    _assin(ws, linha+1)
    ws.column_dimensions["A"].width=8; ws.column_dimensions["B"].width=50
    ws.column_dimensions["C"].width=22; ws.column_dimensions["D"].width=18

def criar_comparativo(wb, df):
    ws = wb.create_sheet("3. CDI/IMAB")
    ws.cell(row=1, column=1, value=f"{NOME_FERRAMENTA} — FIDCs vs CDI vs IMAB").font = FONT_T
    ws.cell(row=2, column=1, value=f"Fontes: CVM | ANBIMA Data | {AUTOR}").font = FONT_S
    ws.merge_cells("A1:H1"); ws.merge_cells("A2:H2")
    CDIa = BENCH["CDI"].get("2026",10.50); IMAa = BENCH["IMAB"].get("2026",11.80)
    cab = ["Fundo","Rentabilidade",f"CDI ({CDIa}%)",f"IMAB ({IMAa}%)","vs CDI","vs IMAB","Rating","Classe"]
    linha = 4
    for c, t in enumerate(cab, 1): ws.cell(row=linha, column=c, value=t)
    _cab(ws, linha, len(cab)); linha += 1
    if 'RENTABILIDADE' in df.columns:
        for _, r in df.sort_values('RENTABILIDADE', ascending=False).head(20).iterrows():
            rent = r.get('RENTABILIDADE',0)
            if pd.isna(rent): continue
            ws.cell(row=linha, column=1, value=str(r.get('DENOMINACAO_SOCIAL',''))[:45])
            ws.cell(row=linha, column=2, value=round(rent,2)).alignment = AR
            ws.cell(row=linha, column=3, value=CDIa).alignment = AR
            ws.cell(row=linha, column=4, value=IMAa).alignment = AR
            vc = rent-CDIa; vi = rent-IMAa
            c1 = ws.cell(row=linha, column=5, value=round(vc,2)); c1.alignment=AR
            c1.font=Font(name="Calibri",bold=True,color="006100") if vc>0 else Font(name="Calibri",bold=True,color="9C0006")
            c1.fill=FILL_V if vc>0 else FILL_R
            c2 = ws.cell(row=linha, column=6, value=round(vi,2)); c2.alignment=AR
            c2.font=Font(name="Calibri",bold=True,color="006100") if vi>0 else Font(name="Calibri",bold=True,color="9C0006")
            c2.fill=FILL_V if vi>0 else FILL_R
            ws.cell(row=linha, column=7, value=r.get('RATING',''))
            ws.cell(row=linha, column=8, value=r.get('CLASSE',''))
            linha += 1
    linha += 3; ws.cell(row=linha, column=1, value="Benchmarks Históricos").font = FONT_T2; linha+=1
    for c,t in enumerate(["Ano","CDI","IMAB","IPCA","CDI Real","IMAB Real"],1): ws.cell(row=linha,column=c,value=t)
    _cab(ws, linha, 6); linha+=1
    for ano in sorted(BENCH["CDI"]):
        cdi=BENCH["CDI"][ano]; imab=BENCH["IMAB"][ano]; ipca=BENCH["IPCA"][ano]
        ws.cell(row=linha,column=1,value=ano); ws.cell(row=linha,column=2,value=cdi).alignment=AR
        ws.cell(row=linha,column=3,value=imab).alignment=AR; ws.cell(row=linha,column=4,value=ipca).alignment=AR
        ws.cell(row=linha,column=5,value=round((1+cdi/100)/(1+ipca/100)-1,4)).alignment=AR
        ws.cell(row=linha,column=6,value=round((1+imab/100)/(1+ipca/100)-1,4)).alignment=AR; linha+=1
    linha+=2; ws.cell(row=linha,column=1,value="⚠ Comparação ilustrativa. Riscos distintos.").font=FONT_S; _assin(ws,linha+1)
    for c,w in [(1,40),(2,16),(3,14),(4,14),(5,14),(6,14),(7,10),(8,16)]:
        ws.column_dimensions[ws.cell(row=1,column=c).column_letter].width = w
    ws.freeze_panes = "A5"

def criar_governanca(wb, metricas):
    ws = wb.create_sheet("4. Governança")
    MAPA = {'VL_PL':'PL (R$ milhões)','PRAZO_MEDIO':'Prazo Médio (dias)','PDD_PCT':'PDD (% PL)',
            'RECOMPRA_PCT':'Recompra (%)','CONC_TOP5_SACADOS':'Conc. Top 5 Sacados (%)',
            'CONC_TOP5_CEDENTES':'Conc. Top 5 Cedentes (%)','NUM_SACADOS':'Nº Sacados',
            'NUM_CEDENTES':'Nº Cedentes','OVERCOLLATERALIZATION':'Overcollateralization (x)',
            'PCT_LIQUIDEZ':'Liquidez (%)','PCT_VENCIDOS':'Vencidos +90d (%)','RENTABILIDADE':'Rentabilidade (% a.a.)'}
    INTERP = {'VL_PL':'Fundos < R$ 50 MM requerem atenção.','PRAZO_MEDIO':'Acima de 180d: risco de alongamento.',
              'PDD_PCT':'🚨 Até 2%: ok. 2-5%: atenção. >5%: alerta.',
              'RECOMPRA_PCT':'>80%: confiança do cedente.','CONC_TOP5_SACADOS':'🚨 >60%: risco severo.',
              'CONC_TOP5_CEDENTES':'🚨 >70%: dependência crítica.','NUM_SACADOS':'Mais pulverizado = menos risco.',
              'NUM_CEDENTES':'Diversificação reduz risco.','OVERCOLLATERALIZATION':'🛡️ >1.2x: saudável. <1.0x: crítico.',
              'PCT_LIQUIDEZ':'Ideal >10%. <5%: risco.','PCT_VENCIDOS':'🚨 Até 2%: normal. >5%: problema.',
              'RENTABILIDADE':'Analisar com PDD e prazo médio.'}
    cab = ["Métrica","Média","Mediana","Desv.Padrão","Melhor","Pior","P25","P75","Interpretação"]
    for c,t in enumerate(cab,1): ws.cell(row=1,column=c,value=t)
    _cab(ws,1,len(cab)); linha=2
    for met,vals in metricas.items():
        nome=MAPA.get(met,met); ws.cell(row=linha,column=1,value=nome).font=Font(name="Calibri",bold=True)
        for ci,ch in enumerate(['media','mediana','desvio_padrao','min','max','p25','p75'],2):
            ws.cell(row=linha,column=ci,value=vals.get(ch)).alignment=AR
        ws.cell(row=linha,column=9,value=INTERP.get(met,''))
        if met in ('PDD_PCT','CONC_TOP5_SACADOS','CONC_TOP5_CEDENTES','PCT_VENCIDOS'):
            md=vals.get('media',0)
            if md>5: ws.cell(row=linha,column=2).fill=FILL_R
            elif md>3: ws.cell(row=linha,column=2).fill=FILL_A
        linha+=1
    _assin(ws,linha+2)
    for c,w in [(1,28),(2,12),(3,12),(4,14),(5,12),(6,12),(7,10),(8,10),(9,70)]:
        ws.column_dimensions[ws.cell(row=1,column=c).column_letter].width = w
    ws.freeze_panes = "A2"

def criar_proposito(wb):
    ws = wb.create_sheet("5. Propósito")
    textos = [
        (f"O Propósito do {NOME_FERRAMENTA}", FONT_T),
        (f"Por {AUTOR} — {VERSAO}", FONT_S),
        ("", None),
        ("O que você está prestes a ver não é uma planilha. É um mapa de minas terrestres.", FONT_D),
        ("", None),
        ("Se existe uma verdade que Morgan Housel ensinou, é esta:\no maior risco não está no que você consegue ver — está no que você não consegue.\n\nVocê olha para a rentabilidade de um FIDC e vê 13% ao ano. Lindo. Mas o que está por trás?\nQual o prazo médio? Qual a concentração? Qual a PDD?\n\nUm fundo pode quebrar não porque rendeu pouco, mas porque ninguém olhou para o lastro.", None),
        ("", None),
        ("🏛️ A 'Basileia' dos FIDCs", FONT_D),
        ("📌 PILAR 1 — CAPITAL: Subordinação + Overcollateralization\n   → O colchão que protege o investidor sênior. Equivalente aos 8% de capital mínimo dos bancos.\n\n📌 PILAR 2 — RISCO: PDD + Concentração + Prazo Médio + Liquidez\n   → A gestão ativa do risco de crédito.\n\n📌 PILAR 3 — TRANSPARÊNCIA: Informe Mensal CVM + Rating\n   → A prestação de contas pública.\n\nDiferença? Bancos são obrigados por lei a seguir Basileia. FIDCs não.\nQuando você exige esses números, você aplica padrão regulatório onde o regulador não chegou.", None),
        ("", None),
        ("⚠️ A analogia tem limites:", FONT_A),
        ("• Dados autodeclarados, não auditados\n• Cada classe de FIDC tem realidades diferentes\n• Defasagem de 30-45 dias\n• Nem todo fundo preenche todos os campos\n\nUse como triagem, não como verdade absoluta.", Font(name="Calibri", size=10, color=CINZA_T)),
        ("", None),
        ("Como usar:", FONT_D),
        ("1. Abra o Ranking e filtre por classe\n2. Veja o Top 10 em cada métrica\n3. Compare com CDI e IMAB\n4. Interprete na aba Governança\n\nPergunte: \"Se esse fundo estivesse no meu portfólio, eu dormiria bem?\"", None),
        ("", None),
        (f"📬 {AUTOR}\n{REPOSITORIO}", FONT_S),
    ]
    linha = 1
    for txt, st in textos:
        if txt == "": linha+=1; continue
        c = ws.cell(row=linha, column=1, value=txt)
        if st: c.font = st
        c.alignment = AL
        nl = max(len(str(txt).split('\n')), len(str(txt))//85+1)
        ws.row_dimensions[linha].height = max(nl*16, 20)
        linha += 1
    linha += 2
    c = ws.cell(row=linha, column=1,
                value='"O segredo do bom investimento não é saber o que vai acontecer. É aceitar que você não sabe, e se preparar para qualquer coisa."\n— Morgan Housel, A Psicologia Financeira')
    c.font = Font(name="Calibri", italic=True, size=11, color=CINZA_T); c.alignment = Alignment(wrap_text=True)
    ws.column_dimensions["A"].width = 120

def criar_fontes(wb):
    ws = wb.create_sheet("6. Fontes")
    linhas = [
        ("FONTES OFICIAIS","t2"),
        ("",""),
        ("1. CVM — Portal Dados Abertos","s"),
        ("   https://dados.cvm.gov.br/dataset/fidc-doc-inf_mensal","l"),
        ("   Informe Mensal de FIDCs — PL, carteira, PDD, prazo médio, recompra","d"),
        ("",""),
        ("2. ANBIMA Data","s"),
        ("   https://data.anbima.com.br","l"),
        ("   Rankings de gestores, dashboard de FIDCs, API RCVM 175","d"),
        ("",""),
        ("3. Benchmarks CDI/IMAB/IPCA","s"),
        ("   Fonte: ANBIMA / B3","d"),
        ("",""),
        ("PROCESSAMENTO","t2"),
        ("",""),
        ("  1. Script acessa repositório CVM e lista arquivos","p"),
        ("  2. Baixa ZIP mais recente (inf_mensal_fidc_YYYYMM.zip)","p"),
        ("  3. Extrai CSVs — cada tabela é um arquivo (TAB_I a TAB_IX)","p"),
        ("  4. Filtra fundos com CLASSE = Duplicatas ou PME","p"),
        ("  5. Calcula métricas derivadas (PDD/PL, overcollateralization, etc.)","p"),
        ("  6. Gera o dashboard Excel com formatação condicional","p"),
        ("",""),
        ("⚠ LIMITAÇÕES","av"),
        ("",""),
        ("  • Formato dos CSVs pode mudar sem aviso (já aconteceu em 2023/2024)","d"),
        ("  • PDD, prazo médio e recompra frequentemente vêm zerados","d"),
        ("  • Classificação Duplicatas/PME é aproximada","d"),
        ("  • Defasagem de 30-45 dias","d"),
        ("",""),
        (f"{NOME_FERRAMENTA} — PÚBLICO E GRATUITO","dest"),
        (f"Autor: {AUTOR}",""),
        (f"Repositório: {REPOSITORIO}",""),
        ("Licença: MIT",""),
    ]
    est = {"t2":Font(name="Calibri",bold=True,size=14,color=AZUL_ESCURO),"s":Font(name="Calibri",bold=True,size=12,color=AZUL_MEDIO),
           "l":Font(name="Calibri",size=10,color="0563C1",underline="single"),"d":Font(name="Calibri",size=10,color=PRETO),
           "p":Font(name="Calibri",size=10),"av":Font(name="Calibri",bold=True,size=10,color="CC0000"),
           "dest":Font(name="Calibri",bold=True,size=11,color=AZUL_ESCURO)}
    for r,(txt,st) in enumerate(linhas,1):
        c=ws.cell(row=r,column=1,value=txt)
        if st in est: c.font=est[st]
    ws.column_dimensions["A"].width=110

def criar_glossario(wb):
    ws = wb.create_sheet("7. Glossário")
    dados = [
        ("Métrica","O que é","Como interpretar"),
        ("PL (R$ milhões)","Patrimônio Líquido do fundo.",">R$200 MM: líquido. <R$50 MM: atenção."),
        ("Prazo Médio (dias)","Prazo médio ponderado dos direitos creditórios.","Até 90d: curto. 90-180: médio. >180: risco."),
        ("PDD (% PL)","Provisão para Devedores Duvidosos.","Até 2%: saudável. 2-5%: atenção. >5%: alerta."),
        ("Recompra (%)","Direitos recomprados no mês / PL.",">80%: confiança. <50%: atenção."),
        ("Overcollateralization","Ativo total / PL.",">1.2x: bom. 1.0-1.2x: apertado. <1.0x: crítico."),
        ("Conc. Top 5 Sacados","% nos 5 maiores devedores.","Até 40%: bom. 40-60%: moderado. >60%: severo."),
        ("Conc. Top 5 Cedentes","% nos 5 maiores cedentes.","Até 50%: saudável. >70%: dependência crítica."),
        ("Liquidez (%)","Caixa + títulos públicos / PL.","Ideal >10%. <5%: risco."),
        ("Vencidos +90d (%)","Créditos com atraso >90d / PL.","Até 2%: normal. 2-5%: investigar. >5%: problema."),
        ("Subordinado (%)","Cotas que absorvem perdas primeiro.","Maior = mais proteção ao sênior. Mínimo: 10%."),
        ("CDI","Certificado de Depósito Interbancário.","Principal benchmark de RF."),
        ("IMAB","Índice ANBIMA de títulos prefixados.","Benchmark para FIDCs de duration longa."),
    ]
    for r, row in enumerate(dados, 1):
        for c, v in enumerate(row, 1): ws.cell(row=r, column=c, value=v)
    _cab(ws,1,3); _zeb(ws,2,len(dados),3)
    ws.column_dimensions["A"].width=28; ws.column_dimensions["B"].width=48
    ws.column_dimensions["C"].width=60; ws.freeze_panes="A2"

def gerar_dashboard(df_fundos, metricas, output_path):
    wb = Workbook()
    criar_capa(wb)
    criar_ranking(wb, df_fundos)
    criar_top10(wb, df_fundos)
    criar_comparativo(wb, df_fundos)
    criar_governanca(wb, metricas)
    criar_proposito(wb)
    criar_fontes(wb)
    criar_glossario(wb)
    wb.save(output_path)
    print(f"\n✅ {NOME_FERRAMENTA} gerado com sucesso!")
    print(f"📁 {output_path}")
    print(f"📋 8 abas: Capa | Ranking | Top10 | CDI/IMAB | Governança | Propósito | Fontes | Glossário")
